import csv
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounting.models import AssetSnapshot, Budget, Category, CategoryType, Entry


# Shared styles
_navy = "1E3A5F"
_header_font = Font(bold=True, color="FFFFFF", size=10)
_header_fill = PatternFill(start_color=_navy, end_color=_navy, fill_type="solid")
_header_align = Alignment(horizontal="center", vertical="center")
_money_fmt = '#,##0.00 "€"'
_pct_fmt = '0"%"'
_thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
_section_income_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
_section_expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_totals_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
_totals_font = Font(bold=True, size=10)
_balance_fill = PatternFill(start_color=_navy, end_color=_navy, fill_type="solid")
_balance_font = Font(bold=True, color="FFFFFF", size=11)
_highlight_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")


def _style_header_row(ws, col_count):
    """Apply header styling to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border


def _apply_border_range(ws, min_row, max_row, max_col):
    """Apply thin borders to a range of cells."""
    for row in range(min_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = _thin_border


def _auto_width(ws, col_count, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col in range(1, col_count + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, max_width)


def generate_journal_excel(fiscal_year):
    """Generate a styled Excel workbook of the journal."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    org = fiscal_year.organization

    # Title rows
    ws.append([f"Journal comptable — {org.name}"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14, color=_navy)
    ws.append([f"Exercice {fiscal_year.start_date.strftime('%d/%m/%Y')} → {fiscal_year.end_date.strftime('%d/%m/%Y')}"])
    ws.merge_cells("A2:E2")
    ws["A2"].font = Font(size=10, color="64748B")
    ws.append([])

    # Headers
    headers = ["Date", "Description", "Catégorie", "Recette", "Dépense"]
    ws.append(headers)
    header_row = 4
    for col in range(1, 6):
        cell = ws.cell(row=header_row, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    entries = Entry.objects.filter(fiscal_year=fiscal_year).select_related("category").order_by("date", "pk")

    total_income = Decimal("0")
    total_expense = Decimal("0")
    row_num = header_row
    for entry in entries:
        income = entry.amount if entry.entry_type == CategoryType.INCOME else None
        expense = entry.amount if entry.entry_type == CategoryType.EXPENSE else None
        if income:
            total_income += income
        if expense:
            total_expense += expense
        row_num += 1
        ws.append([
            entry.date.strftime("%d/%m/%Y"),
            entry.description,
            entry.category.name,
            float(income) if income else "",
            float(expense) if expense else "",
        ])
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = _thin_border
        ws.cell(row=row_num, column=4).number_format = _money_fmt
        ws.cell(row=row_num, column=5).number_format = _money_fmt
        if row_num % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Totals
    row_num += 1
    ws.append(["", "", "TOTAUX", float(total_income), float(total_expense)])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    ws.cell(row=row_num, column=4).number_format = _money_fmt
    ws.cell(row=row_num, column=5).number_format = _money_fmt

    # Balance
    row_num += 1
    ws.append(["", "", "SOLDE", float(total_income - total_expense)])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _balance_font
        cell.fill = _balance_fill
        cell.border = _thin_border
    ws.cell(row=row_num, column=4).number_format = _money_fmt

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_journal_csv(fiscal_year):
    """Generate a CSV string of the journal for a given fiscal year."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    headers = ["Date", "Description", "Catégorie", "Recette", "Dépense"]
    writer.writerow(headers)

    entries = Entry.objects.filter(fiscal_year=fiscal_year).select_related("category").order_by("date", "pk")

    total_income = Decimal("0")
    total_expense = Decimal("0")
    for entry in entries:
        income = entry.amount if entry.entry_type == CategoryType.INCOME else None
        expense = entry.amount if entry.entry_type == CategoryType.EXPENSE else None
        if income:
            total_income += income
        if expense:
            total_expense += expense
        writer.writerow([
            entry.date.strftime("%d/%m/%Y"),
            entry.description,
            entry.category.name,
            str(income) if income else "",
            str(expense) if expense else "",
        ])

    writer.writerow([])
    writer.writerow(["", "", "TOTAUX", str(total_income), str(total_expense)])
    writer.writerow(["", "", "SOLDE", str(total_income - total_expense)])

    return output.getvalue()


def generate_budget_tracking_excel(fiscal_year):
    """Generate a styled Excel workbook comparing budget vs actual."""
    from django.db.models import Sum

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget vs Réalisé"
    org = fiscal_year.organization
    categories = Category.objects.filter(organization=org).order_by("category_type", "name")

    # Title
    ws.append([f"Budget / Réalisé — {org.name}"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14, color=_navy)
    ws.append([f"Exercice {fiscal_year.start_date.strftime('%d/%m/%Y')} → {fiscal_year.end_date.strftime('%d/%m/%Y')}"])
    ws.merge_cells("A2:E2")
    ws["A2"].font = Font(size=10, color="64748B")
    ws.append([])

    # Headers
    headers = ["Catégorie", "Budget", "Réalisé", "Écart", "%"]
    ws.append(headers)
    header_row = 4
    for col in range(1, 6):
        cell = ws.cell(row=header_row, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    row_num = header_row
    total_budget_income = Decimal("0")
    total_actual_income = Decimal("0")
    total_budget_expense = Decimal("0")
    total_actual_expense = Decimal("0")

    # Section: RECETTES
    row_num += 1
    ws.append(["RECETTES", "", "", "", ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = _section_income_fill
        cell.border = _thin_border

    for cat in categories:
        if cat.category_type != CategoryType.INCOME:
            continue
        actual = Entry.objects.filter(fiscal_year=fiscal_year, category=cat).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        budget = Budget.objects.filter(fiscal_year=fiscal_year, category=cat).first()
        planned = budget.planned_amount if budget else Decimal("0")
        diff = actual - planned
        pct = float(actual / planned * 100) if planned else None
        row_num += 1
        ws.append([cat.name, float(planned), float(actual), float(diff), pct or ""])
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = _thin_border
        ws.cell(row=row_num, column=2).number_format = _money_fmt
        ws.cell(row=row_num, column=3).number_format = _money_fmt
        ws.cell(row=row_num, column=4).number_format = _money_fmt
        if pct is not None:
            ws.cell(row=row_num, column=5).number_format = '0"%"'
        if pct is not None and pct < 50:
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = _highlight_fill
        total_budget_income += planned
        total_actual_income += actual

    # Total recettes
    row_num += 1
    ws.append(["Total recettes", float(total_budget_income), float(total_actual_income), float(total_actual_income - total_budget_income), ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    for col in [2, 3, 4]:
        ws.cell(row=row_num, column=col).number_format = _money_fmt

    # Empty row
    row_num += 1
    ws.append([])

    # Section: DÉPENSES
    row_num += 1
    ws.append(["DÉPENSES", "", "", "", ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = _section_expense_fill
        cell.border = _thin_border

    for cat in categories:
        if cat.category_type != CategoryType.EXPENSE:
            continue
        actual = Entry.objects.filter(fiscal_year=fiscal_year, category=cat).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        budget = Budget.objects.filter(fiscal_year=fiscal_year, category=cat).first()
        planned = budget.planned_amount if budget else Decimal("0")
        diff = actual - planned
        pct = float(actual / planned * 100) if planned else None
        row_num += 1
        ws.append([cat.name, float(planned), float(actual), float(diff), pct or ""])
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = _thin_border
        ws.cell(row=row_num, column=2).number_format = _money_fmt
        ws.cell(row=row_num, column=3).number_format = _money_fmt
        ws.cell(row=row_num, column=4).number_format = _money_fmt
        if pct is not None:
            ws.cell(row=row_num, column=5).number_format = '0"%"'
        if pct is not None and pct > 100:
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = _highlight_fill
        total_budget_expense += planned
        total_actual_expense += actual

    # Total dépenses
    row_num += 1
    ws.append(["Total dépenses", float(total_budget_expense), float(total_actual_expense), float(total_actual_expense - total_budget_expense), ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    for col in [2, 3, 4]:
        ws.cell(row=row_num, column=col).number_format = _money_fmt

    # Empty row
    row_num += 1
    ws.append([])

    # Balance
    actual_balance = total_actual_income - total_actual_expense
    budget_balance = total_budget_income - total_budget_expense
    row_num += 1
    ws.append(["SOLDE (Recettes - Dépenses)", float(budget_balance), float(actual_balance), float(actual_balance - budget_balance), ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _balance_font
        cell.fill = _balance_fill
        cell.border = _thin_border
    for col in [2, 3, 4]:
        ws.cell(row=row_num, column=col).number_format = _money_fmt

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_annual_accounts_excel(fiscal_year):
    """Generate a styled Excel workbook for the annual accounts."""
    from django.db.models import Sum

    wb = Workbook()
    org = fiscal_year.organization

    # ===== Sheet 1: Summary =====
    ws = wb.active
    ws.title = "Comptes annuels"

    ws.append([f"Comptes annuels — {org.name}"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=14, color=_navy)
    ws.append([f"Exercice {fiscal_year.start_date.strftime('%d/%m/%Y')} → {fiscal_year.end_date.strftime('%d/%m/%Y')}"])
    ws.merge_cells("A2:B2")
    ws["A2"].font = Font(size=10, color="64748B")
    ws.append([])

    # Headers
    ws.append(["Catégorie", "Montant"])
    header_row = 4
    for col in range(1, 3):
        cell = ws.cell(row=header_row, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    categories = Category.objects.filter(organization=org).order_by("category_type", "name")
    total_income = Decimal("0")
    total_expense = Decimal("0")
    row_num = header_row

    # RECETTES section
    row_num += 1
    ws.append(["RECETTES", ""])
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = _section_income_fill
        cell.border = _thin_border

    for cat in categories:
        if cat.category_type != CategoryType.INCOME:
            continue
        total = Entry.objects.filter(fiscal_year=fiscal_year, category=cat).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        if total > 0:
            row_num += 1
            ws.append([cat.name, float(total)])
            ws.cell(row=row_num, column=1).border = _thin_border
            ws.cell(row=row_num, column=2).border = _thin_border
            ws.cell(row=row_num, column=2).number_format = _money_fmt
            total_income += total

    row_num += 1
    ws.append(["TOTAL RECETTES", float(total_income)])
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    ws.cell(row=row_num, column=2).number_format = _money_fmt

    row_num += 1
    ws.append([])

    # DÉPENSES section
    row_num += 1
    ws.append(["DÉPENSES", ""])
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = _section_expense_fill
        cell.border = _thin_border

    for cat in categories:
        if cat.category_type != CategoryType.EXPENSE:
            continue
        total = Entry.objects.filter(fiscal_year=fiscal_year, category=cat).aggregate(total=Sum("amount"))["total"] or Decimal("0")
        if total > 0:
            row_num += 1
            ws.append([cat.name, float(total)])
            ws.cell(row=row_num, column=1).border = _thin_border
            ws.cell(row=row_num, column=2).border = _thin_border
            ws.cell(row=row_num, column=2).number_format = _money_fmt
            total_expense += total

    row_num += 1
    ws.append(["TOTAL DÉPENSES", float(total_expense)])
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    ws.cell(row=row_num, column=2).number_format = _money_fmt

    row_num += 1
    ws.append([])

    # RÉSULTAT
    row_num += 1
    ws.append(["RÉSULTAT DE L'EXERCICE", float(total_income - total_expense)])
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _balance_font
        cell.fill = _balance_fill
        cell.border = _thin_border
    ws.cell(row=row_num, column=2).number_format = _money_fmt

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # ===== Sheet 2: Journal =====
    ws_journal = wb.create_sheet("Journal")
    ws_journal.append(["Date", "Description", "Catégorie", "Recette", "Dépense"])
    for col in range(1, 6):
        cell = ws_journal.cell(row=1, column=col)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    entries = Entry.objects.filter(fiscal_year=fiscal_year).select_related("category").order_by("date", "pk")
    row_num = 1
    for entry in entries:
        income = float(entry.amount) if entry.entry_type == CategoryType.INCOME else ""
        expense = float(entry.amount) if entry.entry_type == CategoryType.EXPENSE else ""
        row_num += 1
        ws_journal.append([entry.date.strftime("%d/%m/%Y"), entry.description, entry.category.name, income, expense])
        for col in range(1, 6):
            ws_journal.cell(row=row_num, column=col).border = _thin_border
        ws_journal.cell(row=row_num, column=4).number_format = _money_fmt
        ws_journal.cell(row=row_num, column=5).number_format = _money_fmt

    # Totals
    row_num += 1
    ws_journal.append(["", "", "TOTAUX", float(total_income), float(total_expense)])
    for col in range(1, 6):
        cell = ws_journal.cell(row=row_num, column=col)
        cell.font = _totals_font
        cell.fill = _totals_fill
        cell.border = _thin_border
    ws_journal.cell(row=row_num, column=4).number_format = _money_fmt
    ws_journal.cell(row=row_num, column=5).number_format = _money_fmt

    row_num += 1
    ws_journal.append(["", "", "SOLDE", float(total_income - total_expense)])
    for col in range(1, 6):
        cell = ws_journal.cell(row=row_num, column=col)
        cell.font = _balance_font
        cell.fill = _balance_fill
        cell.border = _thin_border
    ws_journal.cell(row=row_num, column=4).number_format = _money_fmt

    ws_journal.column_dimensions["A"].width = 14
    ws_journal.column_dimensions["B"].width = 40
    ws_journal.column_dimensions["C"].width = 22
    ws_journal.column_dimensions["D"].width = 16
    ws_journal.column_dimensions["E"].width = 16

    # ===== Sheet 3: Patrimoine =====
    ws_pat = wb.create_sheet("Patrimoine")
    ws_pat.append(["État du patrimoine"])
    ws_pat["A1"].font = Font(bold=True, size=14, color=_navy)

    snapshot = AssetSnapshot.objects.filter(fiscal_year=fiscal_year).order_by("-date").first()
    if snapshot:
        ws_pat.append([f"Relevé au {snapshot.date.strftime('%d/%m/%Y')}"])
        ws_pat["A2"].font = Font(size=10, color="64748B")
        ws_pat.append([])

        ws_pat.append(["Poste", "Montant"])
        for col in range(1, 3):
            cell = ws_pat.cell(row=4, column=col)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = _header_align
            cell.border = _thin_border

        # Actifs
        ws_pat.append(["ACTIFS", ""])
        for col in range(1, 3):
            cell = ws_pat.cell(row=5, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = _section_income_fill
            cell.border = _thin_border

        items = [("Caisse", snapshot.cash), ("Banque", snapshot.bank), ("Créances", snapshot.receivables)]
        for i, (label, val) in enumerate(items):
            row = 6 + i
            ws_pat.append([label, float(val)])
            ws_pat.cell(row=row, column=1).border = _thin_border
            ws_pat.cell(row=row, column=2).border = _thin_border
            ws_pat.cell(row=row, column=2).number_format = _money_fmt

        # Passifs
        ws_pat.append(["PASSIFS", ""])
        row = 9
        for col in range(1, 3):
            cell = ws_pat.cell(row=row, column=col)
            cell.font = Font(bold=True, size=10)
            cell.fill = _section_expense_fill
            cell.border = _thin_border

        ws_pat.append(["Dettes", float(snapshot.debts)])
        ws_pat.cell(row=10, column=1).border = _thin_border
        ws_pat.cell(row=10, column=2).border = _thin_border
        ws_pat.cell(row=10, column=2).number_format = _money_fmt

        ws_pat.append([])

        # Valeur nette
        ws_pat.append(["VALEUR NETTE", float(snapshot.net_worth)])
        row = 12
        for col in range(1, 3):
            cell = ws_pat.cell(row=row, column=col)
            cell.font = _balance_font
            cell.fill = _balance_fill
            cell.border = _thin_border
        ws_pat.cell(row=row, column=2).number_format = _money_fmt
    else:
        ws_pat.append(["Aucun relevé disponible"])

    ws_pat.column_dimensions["A"].width = 25
    ws_pat.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
