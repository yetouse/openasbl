import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook

from accounting.models import Category, CategoryType, Entry, FiscalYear
from core.models import Organization
from reports.generators.excel import generate_journal_csv, generate_journal_excel


class ExcelTestMixin:
    def setUp(self):
        self.org = Organization.objects.create(name="Test ASBL", address="Bruxelles")
        self.fy = FiscalYear.objects.create(
            organization=self.org,
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
        )
        self.user = User.objects.create_user("testuser", password="testpass")
        self.cat_income = Category.objects.create(
            organization=self.org, name="Cotisations", category_type=CategoryType.INCOME
        )
        Entry.objects.create(
            fiscal_year=self.fy,
            category=self.cat_income,
            date=date(2025, 3, 15),
            amount=Decimal("100.00"),
            description="Cotisation membre",
            created_by=self.user,
        )


class JournalExcelTest(ExcelTestMixin, TestCase):
    def test_generates_valid_excel(self):
        result = generate_journal_excel(self.fy)
        self.assertIsInstance(result, bytes)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["Date", "Description", "Catégorie", "Recette", "Dépense"])
        # Check data row
        self.assertEqual(ws.cell(row=2, column=2).value, "Cotisation membre")


class JournalCsvTest(ExcelTestMixin, TestCase):
    def test_generates_csv_with_headers_and_data(self):
        result = generate_journal_csv(self.fy)
        self.assertIsInstance(result, str)
        lines = result.strip().split("\n")
        self.assertIn("Date", lines[0])
        self.assertIn("Cotisation membre", lines[1])
        # Semicolon-delimited
        self.assertIn(";", lines[0])
